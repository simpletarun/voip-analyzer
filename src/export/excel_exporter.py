"""Excel report exporter (requires openpyxl)."""

import logging
from typing import Dict

from src.export.base import Exporter
from src.models.session import SessionReport

logger = logging.getLogger(__name__)

HAS_OPENPYXL = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    HAS_OPENPYXL = True
except ImportError:
    pass


class ExcelExporter(Exporter):
    def export(self, report: SessionReport, peers: Dict[str, Dict], path: str) -> bool:
        if not HAS_OPENPYXL:
            logger.error("openpyxl not installed - cannot export Excel")
            return False
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Session"
            ws.append(["Field", "Value"])
            for c in ws[1]:
                c.font = Font(bold=True)
            ws.append(["Timestamp", report.timestamp])
            ws.append(["Duration (s)", f"{report.duration_seconds:.1f}"])
            ws.append(["Total Packets", report.total_packets])
            ws.append(["Total Bytes", report.total_bytes])
            ws.append(["P2P Peers", report.p2p_count])
            ws.append(["Relay Servers", report.relay_count])
            ws.append(["Countries", ", ".join(report.countries)])

            wp = wb.create_sheet("Peers")
            headers = ["IP", "Type", "ISP", "City", "Country", "ASN",
                       "Packets", "Bytes", "Confidence"]
            wp.append(headers)
            for c in wp[1]:
                c.font = Font(bold=True)
            for ip, data in peers.items():
                intel = data.get("intel")
                stats = data.get("stats", {})
                wp.append([
                    ip, data.get("type", "UNKNOWN"),
                    intel.isp if intel else "Unknown",
                    intel.city if intel else "Unknown",
                    intel.country if intel else "Unknown",
                    intel.asn if intel else "N/A",
                    stats.get("packets", 0), stats.get("bytes", 0),
                    f"{intel.confidence:.2f}" if intel else "0.00",
                ])
            wb.save(path)
            logger.info("Excel exported: %s", path)
            return True
        except OSError as exc:
            logger.error("Excel export failed: %s", exc)
            return False
